# importing the openpyxl module to work with Excel Spreadsheets
import openpyxl as xl 
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    # loading the Excel file to app.py
    wb = xl.load_workbook(filename)
    # Accessing "Sheet1" in the "transaction.xlsx" file
    sheet = wb['Sheet1']
    # Accessing the cells in "Sheet1" 
    cell = sheet['a1']
    # Another method for accessing cells in "Sheet1"
    # cell = sheet.cell(1, 1)
    # printing the values in "cell"
    # print(cell.value) # Ouputs "transcation_id", the value in A1
    # total numbers of rows in 
    # print(sheet.max_row)

    # Creating a for-loop generating numbers from 1 to 4
    # Note 1: adding 1 to include the number '4'
    # Starting at 2 to ignore the headers (title)
    for row in range(2, sheet.max_row + 1):
        # accessing the third cell in the sheet, 
        # values in the third column,'price'
        cell = sheet.cell(row, 3)
        # Updating the cell value to (cell.value * 0.9)
        corrected_price = cell.value * 0.9
        # Creating a new variable to store the corrrected price 
        # for every row in the sheet. 
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
        print(cell.value)

    # Using Reference class to to select range of values
    # Specifically row number 2 to 4
    values = Reference(sheet, 
                min_row=2, 
                max_row=sheet.max_row,
                min_col=4, 
                max_col=4)

    # Creating the chart
    # Placing the output in cell number 'e2', right of 'D' column
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # Saving the workbook to a new Excel file,  
    wb.save(filename)



